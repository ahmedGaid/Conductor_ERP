"""Custom fields wired into Item create (twenty-harvest FILE_11): validation on write, integer
MONEY, audit snapshot, XLSX export column."""
from __future__ import annotations

import io

import pytest
from rest_framework.test import APIClient

from erp.audit.models import AuditEntry
from erp.core.custom_fields import create_custom_field_def
from erp.identity.models import User

pytestmark = pytest.mark.django_db


def _admin_client() -> APIClient:
    user = User.objects.create_user(username="item_admin", password="Dev12345!")
    user.is_superuser = True
    user.save(update_fields=["is_superuser"])
    client = APIClient()
    client.force_authenticate(user=user)
    return client


def test_create_item_with_valid_custom_data():
    create_custom_field_def(
        entity_key="inventory.item", key="warranty_months", label_ar="الضمان بالشهور",
        label_en="Warranty (months)", type="NUMBER",
    )
    resp = _admin_client().post(
        "/api/inventory/items",
        {"sku": "SKU-1", "name": "Widget", "custom_data": {"warranty_months": "12"}},
        format="json",
    )
    assert resp.status_code == 201, resp.data
    assert resp.data["data"]["custom_data"] == {"warranty_months": "12"}


def test_create_item_with_invalid_custom_data_is_a_human_blame_free_error():
    create_custom_field_def(
        entity_key="inventory.item", key="grade", label_ar="الفئة", label_en="Grade", type="CHOICE",
        choices=["A", "B"],
    )
    resp = _admin_client().post(
        "/api/inventory/items",
        {"sku": "SKU-2", "name": "Widget", "custom_data": {"grade": "C"}},
        format="json",
    )
    assert resp.status_code == 400
    assert "grade" in resp.data["error"]["data"]["errors"]


def test_create_item_money_custom_field_kept_integer():
    create_custom_field_def(
        entity_key="inventory.item", key="insured_value", label_ar="القيمة المؤمنة",
        label_en="Insured value", type="MONEY",
    )
    resp = _admin_client().post(
        "/api/inventory/items",
        {"sku": "SKU-3", "name": "Widget", "custom_data": {"insured_value": "250000"}},
        format="json",
    )
    assert resp.status_code == 201, resp.data
    assert resp.data["data"]["custom_data"]["insured_value"] == 250000
    assert isinstance(resp.data["data"]["custom_data"]["insured_value"], int)


def test_create_item_records_audit_entry_with_custom_data_snapshot():
    client = _admin_client()
    client.post("/api/inventory/items", {"sku": "SKU-4", "name": "Widget"}, format="json")
    entry = AuditEntry.objects.filter(module="inventory", action="create_item", entity_id="SKU-4").first()
    assert entry is not None
    assert entry.after["custom_data"] == {}


def test_item_export_includes_active_custom_field_column():
    create_custom_field_def(
        entity_key="inventory.item", key="grade", label_ar="الفئة", label_en="Grade", type="TEXT",
    )
    client = _admin_client()
    client.post(
        "/api/inventory/items",
        {"sku": "SKU-5", "name": "Widget", "custom_data": {"grade": "A"}},
        format="json",
    )
    resp = client.get("/api/inventory/items", {"export": "xlsx"})
    assert resp["Content-Type"].startswith("application/vnd.openxmlformats")

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(resp.content))
    values = {cell.value for row in wb.active.iter_rows() for cell in row}
    assert "Grade" in values
    assert "A" in values
