"""Smart Import Engine — streaming file reader (plan session 02).

Turns an uploaded spreadsheet into three things the rest of the engine needs, without ever
loading a whole sheet into memory:

* ``sniff`` — what the file *is*: format (xlsx/csv), sheets + row counts, and, for csv, the
  encoding and delimiter it was actually saved with (Arabic Windows Excel loves cp1256 + ``;``).
* ``read_headers`` — the header row (even when it starts at row 3 under a title/logo block) plus a
  handful of sample rows below it, for the mapping wizard.
* ``iter_rows`` — a lazy ``(row_number, {header: value})`` iterator: the 1M-row path (spec step 25).

Reuses the reader already in the repo — ``openpyxl`` (read-only, values-only) for xlsx and stdlib
``csv`` for csv — so no new dependency enters (index decision 5). Legacy ``.xls`` needs a package we
don't ship, so it is refused with a message key that tells the user to re-save as ``.xlsx``.
"""
from __future__ import annotations

import csv
import io
import re
from collections.abc import Iterator
from dataclasses import dataclass
from pathlib import Path

# --- Magic bytes -----------------------------------------------------------------------------
XLSX_MAGIC = b"PK\x03\x04"  # xlsx is a zip
XLS_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"  # legacy OLE2 compound document

# csv encodings, tried in order. utf-8-sig eats a BOM and is a superset-safe first try; cp1256 is
# the Arabic Windows Excel default and, being single-byte, decodes any bytes — so it is the anchor.
CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp1256")
CSV_DELIMITERS = (",", ";", "\t", "|")

HEADER_SCAN_ROWS = 20  # how deep to look for the header row (title/logo blocks sit above it)
SAMPLE_ROWS = 10  # rows returned below the header for the mapping preview
_PROBE_BYTES = 64 * 1024  # how much we decode to guess encoding/delimiter

# BOM + zero-width + bidi marks stripped from cell text; control chars stripped by regex.
# These ARE the bidi/zero-width marks this cleanup strips, not an attack.
_ZERO_WIDTH = "﻿​‌‍‎‏‪‫‬‭‮"  # nosec B613
_ZW_TABLE = {ord(c): None for c in _ZERO_WIDTH}
_CONTROL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
# Arabic read as cp1256 but decoded as latin-1 turns into a run of Ø/Ù/Ú/Û — the mojibake tell.
_MOJIBAKE_RE = re.compile(r"[ØÙÚÛ]")


class UnsupportedFormat(Exception):
    """Raised for a file we cannot read (legacy .xls, empty). Carries an i18n ``message_key``."""

    def __init__(self, message_key: str):
        self.message_key = message_key
        super().__init__(message_key)


@dataclass
class SheetInfo:
    name: str
    row_count: int


@dataclass
class FileInfo:
    format: str  # "xlsx" | "csv"
    sheets: list[SheetInfo]
    encoding: str | None = None  # csv only
    delimiter: str | None = None  # csv only


@dataclass
class Headers:
    headers: list[str]
    header_row_index: int  # 0-based index of the header within the raw row stream
    samples: list[dict]  # up to SAMPLE_ROWS non-empty rows below the header, {header: value}


# --- Public API ------------------------------------------------------------------------------
def sniff(file) -> FileInfo:
    """Detect format, sheets + row counts, and (csv) the encoding and delimiter."""
    raw = _as_bytes(file)
    if _detect_format(raw) == "xlsx":
        return FileInfo(format="xlsx", sheets=_xlsx_sheets(raw))
    encoding = _csv_encoding(raw)
    delimiter = _csv_delimiter(raw, encoding)
    count = sum(1 for _ in _iter_raw_csv(raw, encoding, delimiter))
    return FileInfo(
        format="csv",
        sheets=[SheetInfo(name="", row_count=count)],
        encoding=encoding,
        delimiter=delimiter,
    )


def read_headers(file, sheet=None) -> Headers:
    """Find the header row + sample rows below it.

    Header = the first row (within the top ``HEADER_SCAN_ROWS``) that is >60% non-empty and holds
    at least one non-numeric text cell — this skips title/logo/blank rows real workbooks stack on
    top. Falls back to the first non-empty row if nothing scores as a header.
    """
    scan: list[list] = []
    for cells in _raw_rows(file, sheet):
        scan.append(cells)
        if len(scan) >= HEADER_SCAN_ROWS + SAMPLE_ROWS * 3:
            break

    # The table's real width — used to reject narrow title/logo rows stacked above the header.
    table_width = max((len(_trim_trailing(r)) for r in scan), default=0)
    header_idx = None
    for i in range(min(HEADER_SCAN_ROWS, len(scan))):
        if _looks_like_header(_trim_trailing(scan[i]), table_width):
            header_idx = i
            break
    if header_idx is None:  # nothing header-shaped — take the first row that has any value
        for i, cells in enumerate(scan):
            if _any_nonempty(cells):
                header_idx = i
                break
    if header_idx is None:
        raise UnsupportedFormat("imports.errors.emptyFile")

    headers = _headerize(_trim_trailing(scan[header_idx]))
    samples: list[dict] = []
    for cells in scan[header_idx + 1:]:
        if not _any_nonempty(cells):
            continue
        samples.append(_row_to_dict(headers, cells))
        if len(samples) >= SAMPLE_ROWS:
            break
    return Headers(headers=headers, header_row_index=header_idx, samples=samples)


def iter_rows(file, sheet=None, start=0) -> Iterator[tuple[int, dict]]:
    """Stream data rows as ``(row_number, {header: raw_value})``.

    ``row_number`` is 1-based over the data rows (header excluded, fully-empty rows skipped) so it
    lines up with ``ImportRow.row_number``. ``start`` skips that many leading data rows — the
    resume anchor for a re-run. Never materializes the sheet: xlsx streams via read-only openpyxl,
    csv via a text-wrapped byte stream.
    """
    header_info = read_headers(file, sheet)
    headers = header_info.headers
    data_start = header_info.header_row_index + 1
    row_number = 0
    for i, cells in enumerate(_raw_rows(file, sheet)):
        if i < data_start or not _any_nonempty(cells):
            continue
        row_number += 1
        if row_number <= start:
            continue
        yield row_number, _row_to_dict(headers, cells)


# --- Format detection + byte access ----------------------------------------------------------
def _as_bytes(file) -> bytes:
    """Accept bytes, a path, or a (seekable) binary file object; always yield raw bytes."""
    if isinstance(file, (bytes, bytearray)):
        return bytes(file)
    if isinstance(file, (str, Path)):
        return Path(file).read_bytes()
    try:
        file.seek(0)  # re-entrant: read_headers + iter_rows each re-open the same file
    except (AttributeError, OSError):
        pass
    data = file.read()
    return data.encode("utf-8") if isinstance(data, str) else data


def _detect_format(raw: bytes) -> str:
    if raw[:8] == XLS_MAGIC:
        raise UnsupportedFormat("imports.errors.xlsFormat")
    if raw[:4] == XLSX_MAGIC:
        return "xlsx"
    if not raw.strip():
        raise UnsupportedFormat("imports.errors.emptyFile")
    return "csv"


def _raw_rows(file, sheet=None) -> Iterator[list]:
    """Cleaned cell lists for every row of the chosen sheet — the shared stream both callers walk."""
    raw = _as_bytes(file)
    if _detect_format(raw) == "xlsx":
        return _iter_raw_xlsx(raw, sheet)
    encoding = _csv_encoding(raw)
    return _iter_raw_csv(raw, encoding, _csv_delimiter(raw, encoding))


# --- xlsx ------------------------------------------------------------------------------------
def _xlsx_sheets(raw: bytes) -> list[SheetInfo]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        return [SheetInfo(name=n, row_count=wb[n].max_row or 0) for n in wb.sheetnames]
    finally:
        wb.close()


def _iter_raw_xlsx(raw: bytes, sheet=None) -> Iterator[list]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb.active
        for row in ws.iter_rows(values_only=True):
            yield [_clean_cell(c) for c in row]
    finally:
        wb.close()


# --- csv -------------------------------------------------------------------------------------
def _csv_encoding(raw: bytes) -> str:
    probe = raw[:_PROBE_BYTES]
    for enc in CSV_ENCODINGS:
        try:
            probe.decode(enc)
            return enc
        except UnicodeDecodeError:
            continue
    return "cp1256"  # single-byte: decodes anything, so this is only reached if the probe is empty


def _csv_delimiter(raw: bytes, encoding: str) -> str:
    sample = raw[:_PROBE_BYTES].decode(encoding, errors="replace")
    lines = [ln for ln in sample.splitlines() if ln.strip()][:20]
    try:
        return csv.Sniffer().sniff("\n".join(lines), delimiters="".join(CSV_DELIMITERS)).delimiter
    except csv.Error:
        first = lines[0] if lines else ""
        counts = {d: first.count(d) for d in CSV_DELIMITERS}
        best = max(counts, key=counts.get)
        return best if counts[best] else ","


def _iter_raw_csv(raw: bytes, encoding: str, delimiter: str) -> Iterator[list]:
    stream = io.TextIOWrapper(io.BytesIO(raw), encoding=encoding, errors="replace", newline="")
    for row in csv.reader(stream, delimiter=delimiter):
        yield [_clean_cell(c) for c in row]


# --- Cell + row helpers ----------------------------------------------------------------------
def _clean_cell(value):
    """Strip BOM/zero-width/control chars and repair Arabic mojibake; leave typed cells alone."""
    if not isinstance(value, str):
        return value  # int / float / datetime / bool / None keep their type
    s = _CONTROL_RE.sub("", value.translate(_ZW_TABLE))
    return _repair_mojibake(s).strip()


def _repair_mojibake(s: str) -> str:
    """cp1256-as-latin1 damage → real Arabic, but only when the reverse actually yields Arabic."""
    if not s or not _MOJIBAKE_RE.search(s):
        return s
    try:
        candidate = s.encode("latin-1").decode("cp1256")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return s
    return candidate if any("؀" <= c <= "ۿ" for c in candidate) else s


def _nonempty(c) -> bool:
    return c is not None and (not isinstance(c, str) or c.strip() != "")


def _any_nonempty(cells) -> bool:
    return any(_nonempty(c) for c in cells)


def _is_number(c) -> bool:
    if isinstance(c, bool):
        return False
    if isinstance(c, (int, float)):
        return True
    if isinstance(c, str):
        t = c.strip().replace(",", "")
        if not t:
            return False
        try:
            float(t)
            return True
        except ValueError:
            return False
    return False


def _looks_like_header(cells, table_width: int) -> bool:
    # A header spans (most of) the table — this is how a narrow title/logo row above it is skipped.
    if len(cells) < 2 or len(cells) < table_width * 0.6:
        return False
    non_empty = [c for c in cells if _nonempty(c)]
    if len(non_empty) / len(cells) <= 0.6:
        return False
    return any(not _is_number(c) for c in non_empty)  # reject a numeric-only row


def _trim_trailing(cells: list) -> list:
    end = len(cells)
    while end > 0 and not _nonempty(cells[end - 1]):
        end -= 1
    return cells[:end]


def _headerize(cells: list) -> list[str]:
    """Header cells → clean, unique string keys (blank → column_N, dupes get a suffix)."""
    out: list[str] = []
    seen: dict[str, int] = {}
    for i, c in enumerate(cells):
        name = "" if c is None else str(c).strip()
        if not name:
            name = f"column_{i + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        out.append(name)
    return out


def _row_to_dict(headers: list[str], cells: list) -> dict:
    return {h: (cells[j] if j < len(cells) else None) for j, h in enumerate(headers)}
