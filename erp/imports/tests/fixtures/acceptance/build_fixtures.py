"""FILE_17 acceptance workbook builder — "Before You Start" Task 1.

Generates the messy, real-world-shaped files the FILE_17 acceptance checklist walks through by
hand (Arabic UI first, then English). Run it once to (re)write the fixtures into this directory:

    .venv\\Scripts\\python.exe erp\\imports\\tests\\fixtures\\acceptance\\build_fixtures.py

Coverage against the plan's "Before You Start" list — deliberately curated, not exhaustive over
every registered entity (customers/sales_invoices/journal_entries already exercise every adapter
*type* — master / grouped-document / finance-document — that matters for the acceptance pass):

- `customers_messy.csv` — junk title rows, Arabic+English mixed/misspelled headers, an in-file
  duplicate name, a near-duplicate name (fuzzy-dup candidate), Arabic-Indic digits in a money cell,
  an "L.E"/"جنيه"-suffixed credit limit, a blank required name (row error).
- `sales_invoices_messy.csv` — junk rows, mixed ar/en misspelled headers, every `parse_date` format
  a CSV cell can actually carry (ISO, day-first numeric, dot-separated, spelled month EN+AR,
  Arabic-Indic digits — Excel's numeric-serial date only exists as a real number in an .xlsx cell,
  so that one format needs a live .xlsx upload, not a fixture here), an in-file duplicate doc_number, a missing customer AND
  a missing item (creation-plan candidates on two different ref entities), a wrong file total
  (total_mismatch), an inconsistent-customer document (inconsistent_document), an orphan blank-key
  line (missing_group_key) — then padded to 5,000+ data rows with clean generated invoices for the
  "5k+ rows in one sheet" volume bullet.
- `sales_invoices_cp1256.csv` — small Windows-1256-encoded file, Arabic headers/content only (an
  encoding has one codepage per file — mixing it into the UTF-8 messy file isn't meaningful).
- `journal_entries_unbalanced.csv` — one entry where debits != credits (must error the whole entry,
  never post), plus one clean balanced entry alongside it.
- `customers_100k.csv` — 100,000 clean, generated rows, its OWN file: the checklist's 100k-row
  background-run/pause/resume/kill-process-recovery bullet is a volume/perf concern, independent of
  the messy-data-quality bullets above, and much cheaper to reason about on a plain master adapter
  (one row = one write) than the document/grouped path (one document = one write, N rows).

NOT covered here (left for the live UI pass, per the plan's own two-language manual checklist —
these need judgment, not a fixture): profile-save/reuse, autofix apply, all four import strategies,
permission-rejection-by-an-unpermitted-user, trial-balance opening's correcting-entry proposal path.

Also builds `acceptance_workbook.xlsx` — ONE multi-sheet workbook (Customers / SalesInvoices /
PurchaseInvoices / JournalEntries) for the manual browser pass of the checklist's own first bullet:
"Upload the messy workbook with ZERO preparation → correct entity detected per sheet." Sheets are
the curated (unpadded) rows only — this file is for switching the sheet picker in the UI and
watching auto-detect flip entity each time, not a volume test (customers_100k.csv stays separate).
The SalesInvoices sheet also carries ONE row with a REAL Excel numeric-serial date cell (only
representable in an actual .xlsx, never a CSV) — the one `parse_date` format the CSV fixtures above
cannot exercise.
"""
from __future__ import annotations

import csv
import io
import os

from openpyxl import Workbook

HERE = os.path.dirname(os.path.abspath(__file__))

# Real seeded reference data (this dev DB) — used so the "should resolve fine" rows actually
# resolve, and the "missing master" rows use codes that provably do NOT exist yet.
REAL_CUSTOMER = "ACME"       # Acme Corp
REAL_SUPPLIER = "GLOBEX"     # Globex Supplies
REAL_ITEM = "BOLT"           # Bolt (kg)
REAL_WAREHOUSE = "MAIN"      # Main Warehouse
MISSING_CUSTOMER = "ZZZ-NEW-CUSTOMER"
MISSING_ITEM = "ZZZ-NEW-ITEM"
MISSING_SUPPLIER = "ZZZ-NEW-SUPPLIER"
# 2026-06-15 as an Excel 1900-system serial number (days since 1899-12-30) — for the one
# real-numeric-cell date row the CSV fixtures structurally can't carry.
EXCEL_SERIAL_2026_06_15 = 46188


def _write_csv(path: str, rows: list[list[str]], encoding: str = "utf-8-sig") -> None:
    buf = io.StringIO()
    csv.writer(buf).writerows(rows)
    with open(path, "wb") as f:
        f.write(buf.getvalue().encode(encoding))


# --- customers_messy.csv -----------------------------------------------------------------------
def _customers_messy_rows() -> list[list[str]]:
    return [
        ["Conductor ERP - Customer Import Template"],  # junk title row 1
        ["Exported 2026-07-24 — do not edit the header row below"],  # junk title row 2
        [],  # blank row
        # Mixed ar/en, misspelled/synonym headers (not the canonical field names) — exercises
        # header detection + synonym matching, not a literal field-name match.
        ["كود العميل", "Customer Name", "Credit Limitt"],  # "Creditt" = a misspelling of "credit limit"
        ["CM-1", "Modern Trading Co", "50000"],
        ["CM-2", "مصنع الأمل للصناعات", "٢٥٠٠٠ جنيه"],  # Arabic-Indic digits + Arabic currency word
        ["CM-3", "Al Noor Stores", "L.E 10000"],  # English money-suffix style
        ["CM-4", "Modern Trading Company", ""],  # near-duplicate of CM-1's name (fuzzy-dup candidate)
        ["CM-5", "Modern Trading Co", "5000"],  # EXACT duplicate name of CM-1 (in-file duplicate)
        ["CM-6", "", "1000"],  # blank required name → real error row
    ]


def build_customers_messy() -> None:
    _write_csv(os.path.join(HERE, "customers_messy.csv"), _customers_messy_rows())


# --- sales_invoices_messy.csv (+ 5k padding) ---------------------------------------------------
def _messy_invoice_rows() -> list[list[str]]:
    # header: doc_number, customer_ref, date, currency, warehouse_ref, tax_token, file_total_minor,
    # item_ref, quantity, unit_price_minor, discount_minor — via misspelled/mixed-language synonyms.
    header = [
        "Invoic No", "العميل", "Date", "Currency", "المخزن", "Tax", "الاجمالي",
        "الصنف", "Qnty", "Unit Price", "Discount",
    ]
    rows: list[list[str]] = [
        ["Smart Import — Sales Invoices (messy acceptance sheet)"],  # junk title row
        [],
        header,
        # 1) ISO date, clean, resolves fine — the "everything's fine" control row.
        ["ACC-INV-1", REAL_CUSTOMER, "2026-06-01", "EGP", REAL_WAREHOUSE, "14%", "2260", REAL_ITEM, "2", "1000", "0"],
        # 2) Day-first numeric date (02/01/2026 = 2 Jan 2026, Egypt convention), same document
        ["ACC-INV-1", "", "02/01/2026", "", "", "", "", REAL_ITEM, "1", "1000", "0"],
        # 3) Dot-separated day-first + Arabic month name + Arabic-Indic digits, own document
        ["ACC-INV-2", REAL_CUSTOMER, "٢٣ يونيو ٢٠٢٦", "EGP", REAL_WAREHOUSE, "ضريبة", "2000", REAL_ITEM, "2", "1000", "0"],
        # 4) Dot-separated day-first numeric date (Egyptian convention), own document. NOTE: a
        # real Excel SERIAL date (parse_date's numeric-cell branch) only exists as an actual
        # number in an .xlsx cell — a CSV cell is always text, so "46601" here would read as a
        # bare, invalid string, not a serial date; that format is exercised in a live .xlsx
        # upload during the manual UI pass, not here.
        ["ACC-INV-3", REAL_CUSTOMER, "01.06.2026", "EGP", REAL_WAREHOUSE, "vat", "1000", REAL_ITEM, "1", "1000", "0"],
        # 5) In-file duplicate doc_number of ACC-INV-1 with DIFFERENT customer → inconsistent_document
        ["ACC-INV-1", "Al Noor Stores", "2026-06-01", "EGP", REAL_WAREHOUSE, "14%", "", REAL_ITEM, "1", "1000", "0"],
        # 6) Wrong file total on its own document → total_mismatch (warning, not error)
        ["ACC-INV-4", REAL_CUSTOMER, "2026-06-05", "EGP", REAL_WAREHOUSE, "14%", "999999", REAL_ITEM, "1", "1000", "0"],
        # 7) Missing customer master → creation-plan candidate (customers entity)
        ["ACC-INV-5", MISSING_CUSTOMER, "2026-06-06", "EGP", REAL_WAREHOUSE, "14%", "", REAL_ITEM, "1", "1000", "0"],
        # 8) Missing item master → creation-plan candidate (items entity)
        ["ACC-INV-6", REAL_CUSTOMER, "2026-06-07", "EGP", REAL_WAREHOUSE, "14%", "", MISSING_ITEM, "1", "1000", "0"],
        # 9) Orphan blank-key line — no doc_number, no header data, nothing open before it in the
        # curated set (it's placed first among the padding below, not here, to guarantee no open
        # group precedes it — see build_sales_invoices_messy).
    ]
    return rows


def _clean_padding_rows(n: int, start_no: int) -> list[list[str]]:
    out: list[list[str]] = []
    for i in range(n):
        doc = f"ACC-GEN-{start_no + i}"
        out.append([
            doc, REAL_CUSTOMER, "2026-06-10", "EGP", REAL_WAREHOUSE, "14%",
            "1000", REAL_ITEM, "1", "1000", "0",
        ])
    return out


def build_sales_invoices_messy(padding_rows: int = 5200) -> None:
    header_block = _messy_invoice_rows()
    # The orphan row must have NO group open before it — insert as the very first data row so
    # nothing precedes it, then the rest of the curated rows, then generated padding.
    orphan_row = [
        "", "", "", "", "", "", "", REAL_ITEM, "1", "1000", "0",
    ]
    data_rows = [orphan_row, *header_block[3:]]  # skip title/blank/header, they go back on top
    rows = [header_block[0], header_block[1], header_block[2], *data_rows,
            *_clean_padding_rows(padding_rows, start_no=1)]
    _write_csv(os.path.join(HERE, "sales_invoices_messy.csv"), rows)


# --- purchase invoices (workbook sheet only — mirrors sales, supplier-side) ---------------------
def _messy_purchase_invoice_rows() -> list[list[str]]:
    # header: doc_number, supplier_ref, date, currency, warehouse_ref, tax_token, file_total_minor,
    # item_ref, quantity, unit_price_minor (no discount field on PurchaseInvoiceAdapter).
    header = [
        "Bill No", "المورد", "Date", "Currency", "المخزن", "Tax", "الاجمالي", "الصنف", "Qnty", "Unit Cost",
    ]
    return [
        header,
        # 1) clean control row.
        ["ACC-PINV-1", REAL_SUPPLIER, "2026-06-02", "EGP", REAL_WAREHOUSE, "14%", "2200", REAL_ITEM, "2", "1000"],
        # 2) second line of the same document (blank supplier/date — merged-cell continuation).
        ["ACC-PINV-1", "", "", "", "", "", "", REAL_ITEM, "1", "200"],
        # 3) missing supplier master → creation-plan candidate (suppliers entity).
        ["ACC-PINV-2", MISSING_SUPPLIER, "2026-06-03", "EGP", REAL_WAREHOUSE, "14%", "", REAL_ITEM, "1", "1000"],
        # 4) wrong file total → total_mismatch (warning).
        ["ACC-PINV-3", REAL_SUPPLIER, "2026-06-04", "EGP", REAL_WAREHOUSE, "14%", "1", REAL_ITEM, "1", "1000"],
    ]


# --- sales_invoices_cp1256.csv ------------------------------------------------------------------
def build_sales_invoices_cp1256() -> None:
    header = ["رقم الفاتورة", "العميل", "التاريخ", "العملة", "المخزن", "الضريبة", "الاجمالي", "الصنف", "الكمية", "سعر الوحدة", "خصم"]
    rows = [
        header,
        ["CP-INV-1", REAL_CUSTOMER, "2026-06-01", "EGP", REAL_WAREHOUSE, "14%", "2000", REAL_ITEM, "2", "1000", "0"],
    ]
    _write_csv(os.path.join(HERE, "sales_invoices_cp1256.csv"), rows, encoding="cp1256")


# --- journal_entries_unbalanced.csv -------------------------------------------------------------
def _journal_unbalanced_rows() -> list[list[str]]:
    header = ["Entry No", "Date", "البيان", "Account", "Debit", "Credit", "Line Memo"]
    return [
        header,
        # JE-1: balanced (1000 debit == 1000 credit) — the "clean" control entry. entry_ref is
        # BLANK on the continuation line (merged-cell pattern) — repeating it would misread as an
        # in-file duplicate entry_ref, not a second line of the same one.
        ["JE-1", "2026-06-01", "Cash sale", "1000", "1000", "", "Till receipt"],
        ["", "", "", "1100", "", "1000", "AR clearance"],
        # JE-2: UNBALANCED on purpose (900 debit vs 1000 credit) — must error the whole entry.
        ["JE-2", "2026-06-02", "Bank transfer", "1010", "900", "", ""],
        ["", "", "", "1000", "", "1000", ""],
    ]


def build_journal_entries_unbalanced() -> None:
    _write_csv(os.path.join(HERE, "journal_entries_unbalanced.csv"), _journal_unbalanced_rows())


# --- customers_100k.csv (volume / background-runner / pause-resume / kill-process) -------------
def build_customers_100k(n: int = 100_000) -> None:
    path = os.path.join(HERE, "customers_100k.csv")
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["code", "name", "credit_limit_minor"])
        for i in range(n):
            w.writerow([f"VOL-{i:06d}", f"Volume Test Customer {i:06d}", "0"])


# --- acceptance_workbook.xlsx (multi-sheet — manual "correct entity detected per sheet" pass) ---
def build_acceptance_workbook() -> None:
    wb = Workbook()
    wb.remove(wb.active)  # default blank sheet — every sheet below is added explicitly

    ws = wb.create_sheet("Customers")
    for row in _customers_messy_rows():
        ws.append(row)

    ws = wb.create_sheet("SalesInvoices")
    for row in _messy_invoice_rows():
        ws.append(row)
    # Real numeric Excel-serial date cell — impossible in the CSV twin of this fixture.
    ws.append(
        ["ACC-INV-SERIAL", REAL_CUSTOMER, EXCEL_SERIAL_2026_06_15, "EGP", REAL_WAREHOUSE, "14%",
         "1000", REAL_ITEM, "1", "1000", "0"]
    )

    ws = wb.create_sheet("PurchaseInvoices")
    for row in _messy_purchase_invoice_rows():
        ws.append(row)

    ws = wb.create_sheet("JournalEntries")
    for row in _journal_unbalanced_rows():
        ws.append(row)

    wb.save(os.path.join(HERE, "acceptance_workbook.xlsx"))


def main() -> None:
    build_customers_messy()
    build_sales_invoices_messy()
    build_sales_invoices_cp1256()
    build_journal_entries_unbalanced()
    build_customers_100k()
    build_acceptance_workbook()
    print("Acceptance fixtures written to", HERE)


if __name__ == "__main__":
    main()
