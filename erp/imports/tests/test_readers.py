"""Streaming file reader (plan session 02).

Fixtures are generated in-process (openpyxl workbooks / encoded csv bytes) rather than committed as
binary blobs — same coverage, no binary churn in the repo. Drift from the plan's "tests/fixtures/"
note is deliberate; intent (each path exercised) wins.
"""
from __future__ import annotations

import io
import types

import pytest
from openpyxl import Workbook

from erp.imports import readers
from erp.imports.readers import FileInfo, Headers, UnsupportedFormat


# --- helpers ---------------------------------------------------------------------------------
def _xlsx(rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _csv(rows, encoding="utf-8", delimiter=",", bom=False) -> bytes:
    text = "\n".join(delimiter.join("" if c is None else str(c) for c in r) for r in rows)
    data = text.encode(encoding)
    return (b"\xef\xbb\xbf" + data) if bom else data


# --- format + sniff --------------------------------------------------------------------------
def test_sniff_xlsx():
    info = readers.sniff(_xlsx([["Name", "Phone"], ["Ali", "0100"]]))
    assert isinstance(info, FileInfo)
    assert info.format == "xlsx"
    assert info.sheets[0].row_count == 2


def test_sniff_csv_utf8_comma():
    info = readers.sniff(_csv([["Name", "Phone"], ["Ali", "0100"], ["Sara", "0111"]]))
    assert info.format == "csv"
    assert info.encoding in ("utf-8-sig", "utf-8")
    assert info.delimiter == ","
    assert info.sheets[0].row_count == 3


def test_sniff_csv_semicolon_delimiter():
    info = readers.sniff(_csv([["Name", "City"], ["Ali", "Cairo"]], delimiter=";"))
    assert info.delimiter == ";"


def test_sniff_csv_cp1256_arabic():
    rows = [["الاسم", "الهاتف"], ["أحمد", "0100"]]
    info = readers.sniff(_csv(rows, encoding="cp1256"))
    assert info.encoding == "cp1256"


# --- headers ---------------------------------------------------------------------------------
def test_headers_simple_csv():
    h = readers.read_headers(_csv([["Name", "Phone"], ["Ali", "0100"], ["Sara", "0111"]]))
    assert isinstance(h, Headers)
    assert h.headers == ["Name", "Phone"]
    assert h.header_row_index == 0
    assert h.samples[0] == {"Name": "Ali", "Phone": "0100"}
    assert len(h.samples) == 2


def test_headers_below_title_rows():
    # Rows 1-2 are a title/subtitle block; the real header is row 3 (index 2).
    rows = [
        ["Customer Master Export", None],
        ["Generated 2026-07-04", None],
        ["Name", "Phone"],
        ["Ali", "0100"],
    ]
    h = readers.read_headers(_xlsx(rows))
    assert h.headers == ["Name", "Phone"]
    assert h.header_row_index == 2
    assert h.samples == [{"Name": "Ali", "Phone": "0100"}]


def test_headers_cp1256_arabic_no_mojibake():
    rows = [["الاسم", "الهاتف"], ["أحمد", "0100"]]
    h = readers.read_headers(_csv(rows, encoding="cp1256"))
    assert h.headers == ["الاسم", "الهاتف"]
    assert h.samples[0]["الاسم"] == "أحمد"


def test_headers_utf8_bom_stripped():
    h = readers.read_headers(_csv([["Name", "Phone"], ["Ali", "0100"]], bom=True))
    assert h.headers == ["Name", "Phone"]  # BOM not glued onto the first header


def test_headers_blank_and_duplicate_columns_made_unique():
    h = readers.read_headers(_csv([["Name", "", "Name"], ["a", "b", "c"]]))
    assert h.headers == ["Name", "column_2", "Name_1"]


# --- streaming rows --------------------------------------------------------------------------
def test_iter_rows_is_lazy_generator():
    gen = readers.iter_rows(_csv([["Name"], ["Ali"]]))
    assert isinstance(gen, types.GeneratorType)


def test_iter_rows_numbers_and_skips_empty():
    rows = [["Name", "Phone"], ["Ali", "0100"], [None, None], ["Sara", "0111"]]
    out = list(readers.iter_rows(_csv(rows)))
    assert out == [
        (1, {"Name": "Ali", "Phone": "0100"}),
        (2, {"Name": "Sara", "Phone": "0111"}),  # blank row skipped, numbering stays contiguous
    ]


def test_iter_rows_start_offset_for_resume():
    rows = [["Name"], ["a"], ["b"], ["c"]]
    out = list(readers.iter_rows(_csv(rows), start=2))
    assert out == [(3, {"Name": "c"})]


def test_iter_rows_10k_streams_without_materializing():
    body = "\n".join(f"row{i}" for i in range(10_000))
    raw = ("Name\n" + body).encode("utf-8")
    count = sum(1 for _ in readers.iter_rows(raw))  # never builds a list
    assert count == 10_000


def test_iter_rows_xlsx_below_title():
    rows = [["Report", None], ["Name", "Phone"], ["Ali", "0100"], ["Sara", "0111"]]
    out = list(readers.iter_rows(_xlsx(rows)))
    assert out == [
        (1, {"Name": "Ali", "Phone": "0100"}),
        (2, {"Name": "Sara", "Phone": "0111"}),
    ]


# --- unsupported / empty ---------------------------------------------------------------------
def test_xls_legacy_rejected_with_message_key():
    with pytest.raises(UnsupportedFormat) as ei:
        readers.sniff(readers.XLS_MAGIC + b"\x00" * 32)
    assert ei.value.message_key == "imports.errors.xlsFormat"


def test_empty_file_rejected():
    with pytest.raises(UnsupportedFormat) as ei:
        readers.sniff(b"")
    assert ei.value.message_key == "imports.errors.emptyFile"
