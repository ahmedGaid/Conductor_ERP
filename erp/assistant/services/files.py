"""Attachment understanding for the chat pipeline (plan session 07).

Turns an uploaded file into model-ready content:
- images / PDF → the provider's vision path (same bytes the extraction flow sends);
- CSV / XLSX   → a compact text table (header + dtypes + first rows + row count);
- JSON / XML / TXT → the first ~8 KB verbatim, marked when truncated.

This session is *understanding and Q&A only* — "what's in this file?", "summarise this price
list". Structured import (create records from a file) is session 12. Nothing here writes business
data; the file text simply rides along with the user's question to the model.
"""
from __future__ import annotations

import csv
import io

from django.http import Http404

from ..models import Attachment
from .extraction import ALLOWED_TYPES as VISION_TYPES  # image/* + application/pdf

CSV_TYPES = {"text/csv", "application/csv", "application/vnd.ms-excel"}
XLSX_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
JSON_TYPES = {"application/json", "text/json"}
XML_TYPES = {"application/xml", "text/xml"}
TXT_TYPES = {"text/plain"}

# Everything the attachment endpoint accepts (allowlist, no sniffing — Session-00 posture).
ALLOWED_ATTACHMENT_TYPES = (
    VISION_TYPES | CSV_TYPES | {XLSX_TYPE} | JSON_TYPES | XML_TYPES | TXT_TYPES
)

# How much of a text/JSON/XML file we hand the model — enough to answer over, bounded for cost.
TEXT_BUDGET = 8 * 1024
# How many data rows of a table we show the model (plus the true total row count).
ROW_SAMPLE = 20

# The tabular types the structured-import pipeline (session 14) reads row by row.
IMPORT_TYPES = CSV_TYPES | {XLSX_TYPE}
# Hard ceiling on rows a single import reads into memory — a runaway file can't exhaust the process.
MAX_IMPORT_ROWS = 5000


def fetch_unclaimed(attachment_ids, *, user) -> list[Attachment]:
    """The caller's own still-unclaimed attachments, in order — or 404 if any id doesn't qualify.

    Used to reject a bad reference *before* the stream opens (so it's a clean 404, not a mid-stream
    error). Any id that is missing, someone else's, or already claimed makes the whole set a 404 —
    the "indistinguishable from absent" posture the rest of the assistant uses.
    """
    try:
        ids = list(dict.fromkeys(int(i) for i in attachment_ids))
    except (TypeError, ValueError):
        raise Http404
    rows = {a.id: a for a in Attachment.objects.filter(id__in=ids, user=user, message__isnull=True)}
    if len(rows) != len(ids):
        raise Http404
    return [rows[i] for i in ids]


def claim_attachments(attachment_ids, *, user, message) -> list[Attachment]:
    """Link the caller's own unclaimed attachments to a message (validates, then sets the FK)."""
    attachments = fetch_unclaimed(attachment_ids, user=user)
    Attachment.objects.filter(id__in=[a.id for a in attachments]).update(message=message)
    for a in attachments:
        a.message = message
    return attachments


def attachment_chips(attachments) -> list[dict]:
    """The compact descriptor stored in the user message's ``meta`` so history re-renders chips."""
    return [
        {"id": a.id, "name": a.name, "content_type": a.content_type, "size": a.size}
        for a in attachments
    ]


def describe_for_model(attachment: Attachment) -> dict:
    """One attachment → model-ready content.

    Returns ``{"media": {"media_type", "data"}}`` for images/PDF (handed to the vision path) or
    ``{"text": "..."}`` for tabular/text files (folded into the user's question text).
    """
    ct = (attachment.content_type or "").lower()
    if ct in VISION_TYPES:
        attachment.file.open("rb")
        try:
            return {"media": {"media_type": ct, "data": attachment.file.read()}}
        finally:
            attachment.file.close()

    raw = _read_bytes(attachment)
    if ct in CSV_TYPES:
        return {"text": _describe_csv(attachment.name, raw)}
    if ct == XLSX_TYPE:
        return {"text": _describe_xlsx(attachment.name, raw)}
    # JSON / XML / TXT — verbatim, truncated to the budget.
    text = raw.decode("utf-8", errors="replace")
    clipped = text[:TEXT_BUDGET]
    note = "\n…(truncated)" if len(text) > TEXT_BUDGET else ""
    return {"text": f'File "{attachment.name}" ({attachment.content_type}):\n{clipped}{note}'}


def read_table(attachment: Attachment) -> tuple[list[str], list[list]]:
    """A tabular attachment → ``(header, data_rows)`` for the structured-import pipeline (session 14).

    Unlike ``describe_for_model`` (which samples a file into model-ready text), this returns the real
    rows so preview/execute can validate and create each one. Header cells are stripped strings; data
    cells stay as read (``None`` for blank XLSX cells) and the parse layer coerces them. Bounded to
    ``MAX_IMPORT_ROWS`` data rows. A non-tabular type returns ``([], [])``.
    """
    ct = (attachment.content_type or "").lower()
    raw = _read_bytes(attachment)
    if ct in CSV_TYPES:
        return _read_csv_rows(raw)
    if ct == XLSX_TYPE:
        return _read_xlsx_rows(raw)
    return [], []


def _read_csv_rows(raw: bytes) -> tuple[list[str], list[list]]:
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    try:
        header = [str(h).strip() for h in next(reader)]
    except StopIteration:
        return [], []
    rows: list[list] = []
    for row in reader:
        if len(rows) >= MAX_IMPORT_ROWS:
            break
        rows.append(list(row))
    return header, rows


def _read_xlsx_rows(raw: bytes) -> tuple[list[str], list[list]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        rows_iter = wb.active.iter_rows(values_only=True)
        try:
            header = [("" if h is None else str(h)).strip() for h in next(rows_iter)]
        except StopIteration:
            return [], []
        rows: list[list] = []
        for row in rows_iter:
            if len(rows) >= MAX_IMPORT_ROWS:
                break
            rows.append(list(row))
        return header, rows
    finally:
        wb.close()


def _read_bytes(attachment: Attachment) -> bytes:
    attachment.file.open("rb")
    try:
        return attachment.file.read()
    finally:
        attachment.file.close()


def _table_text(name: str, header: list[str], rows: list[list[str]], total: int) -> str:
    """A compact pipe table: header + sampled rows + the true total, all as plain text."""
    head = " | ".join(str(h) for h in header) if header else "(no header row)"
    body = "\n".join(" | ".join("" if c is None else str(c) for c in r) for r in rows)
    shown = len(rows)
    more = f"\n…({total - shown} more rows)" if total > shown else ""
    return (
        f'File "{name}" — {len(header)} columns, {total} data rows.\n'
        f"{head}\n{body}{more}"
    )


def _describe_csv(name: str, raw: bytes) -> str:
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    try:
        header = next(reader)
    except StopIteration:
        return f'File "{name}" is empty.'
    sample: list[list[str]] = []
    total = 0
    for row in reader:
        total += 1
        if len(sample) < ROW_SAMPLE:
            sample.append(row)
    return _table_text(name, header, sample, total)


def _describe_xlsx(name: str, raw: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = list(next(rows_iter))
        except StopIteration:
            return f'File "{name}" is empty.'
        sample: list[list] = []
        total = 0
        for row in rows_iter:
            total += 1
            if len(sample) < ROW_SAMPLE:
                sample.append(list(row))
        return _table_text(name, header, sample, total)
    finally:
        wb.close()
