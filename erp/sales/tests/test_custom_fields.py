"""Custom fields wired into Customer create (twenty-harvest FILE_11): validation on write,
integer MONEY, audit snapshot, XLSX export column."""
from __future__ import annotations

import io

import pytest
from rest_framework.test import APIClient

from erp.audit.models import AuditEntry
from erp.core.custom_fields import create_custom_field_def
from erp.identity.models import User

pytestmark = pytest.mark.django_db


def _admin_client() -> APIClient:
    user = User.objects.create_user(username="cust_admin", password="Dev12345!")
    user.is_superuser = True
    user.save(update_fields=["is_superuser"])
    client = APIClient()
    client.force_authenticate(user=user)
    return client


def test_create_customer_with_valid_custom_data():
    create_custom_field_def(
        entity_key="sales.customer", key="zone", label_ar="منطقة", label_en="Zone", type="CHOICE",
        choices=["cairo", "giza"],
    )
    resp = _admin_client().post(
        "/api/sales/customers",
        {"code": "C-1", "name": "Acme", "custom_data": {"zone": "cairo"}},
        format="json",
    )
    assert resp.status_code == 201, resp.data
    assert resp.data["data"]["custom_data"] == {"zone": "cairo"}


def test_create_customer_with_invalid_choice_is_a_human_blame_free_error():
    create_custom_field_def(
        entity_key="sales.customer", key="zone", label_ar="منطقة", label_en="Zone", type="CHOICE",
        choices=["cairo", "giza"],
    )
    resp = _admin_client().post(
        "/api/sales/customers",
        {"code": "C-2", "name": "Acme", "custom_data": {"zone": "alexandria"}},
        format="json",
    )
    assert resp.status_code == 400
    assert "zone" in resp.data["error"]["data"]["errors"]


def test_create_customer_money_custom_field_kept_integer():
    create_custom_field_def(
        entity_key="sales.customer", key="deposit", label_ar="عربون", label_en="Deposit", type="MONEY",
    )
    resp = _admin_client().post(
        "/api/sales/customers",
        {"code": "C-3", "name": "Acme", "custom_data": {"deposit": "5000"}},
        format="json",
    )
    assert resp.status_code == 201, resp.data
    assert resp.data["data"]["custom_data"]["deposit"] == 5000
    assert isinstance(resp.data["data"]["custom_data"]["deposit"], int)


def test_create_customer_records_audit_entry_with_custom_data_snapshot():
    client = _admin_client()
    client.post("/api/sales/customers", {"code": "C-4", "name": "Acme"}, format="json")
    entry = AuditEntry.objects.filter(module="sales", action="create_customer", entity_id="C-4").first()
    assert entry is not None
    assert entry.after["custom_data"] == {}


def test_customer_export_includes_active_custom_field_column():
    create_custom_field_def(
        entity_key="sales.customer", key="zone", label_ar="منطقة", label_en="Zone", type="TEXT",
    )
    client = _admin_client()
    client.post(
        "/api/sales/customers",
        {"code": "C-5", "name": "Acme", "custom_data": {"zone": "cairo"}},
        format="json",
    )
    resp = client.get("/api/sales/customers", {"export": "xlsx"})
    assert resp["Content-Type"].startswith("application/vnd.openxmlformats")

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(resp.content))
    values = {cell.value for row in wb.active.iter_rows() for cell in row}
    assert "Zone" in values
    assert "cairo" in values
