"""Tabular report export — CSV and XLSX renderers + an HttpResponse helper.

A presentation-agnostic ``ReportTable`` (built by each module's API layer) is rendered to bytes.
Money cells carry integer **minor units**; the renderers convert to major units (÷100, 2 decimals).
Arabic is preserved end to end: CSV is UTF-8 **with a BOM** (so Excel detects the encoding) and XLSX
sets a right-to-left sheet view when ``rtl=True``. PDF export is intentionally the browser's native
print-to-PDF (perfect RTL shaping, no bundled fonts) — see the web app's print stylesheet.
"""
from __future__ import annotations

import csv
import datetime as dt
import io
from dataclasses import dataclass, field

from django.http import HttpResponse

EXPORT_FORMATS = ("csv", "xlsx")
_MINOR = 100

_CONTENT_TYPES = {
    "csv": "text/csv; charset=utf-8",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}


@dataclass(frozen=True)
class Column:
    key: str
    label: str
    kind: str = "text"    # text | money | number
    align: str = "start"  # start | end


@dataclass
class ReportTable:
    title: str
    columns: list[Column]
    rows: list[dict]
    footer: list[dict] = field(default_factory=list)            # emphasized total rows
    meta: list[tuple[str, str]] = field(default_factory=list)   # (label, value) context lines
    currency: str = "EGP"
    rtl: bool = False


def _major(value) -> float:
    return round((value or 0) / _MINOR, 2)


def _cell_text(col: Column, row: dict) -> str:
    value = row.get(col.key, "")
    if value is None or value == "":
        return ""
    if col.kind == "money":
        return f"{_major(value):.2f}"
    return str(value)


def to_csv(table: ReportTable) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    if table.title:
        writer.writerow([table.title])
    for label, value in table.meta:
        writer.writerow([label, value])
    if table.title or table.meta:
        writer.writerow([])
    writer.writerow([c.label for c in table.columns])
    for row in (*table.rows, *table.footer):
        writer.writerow([_cell_text(c, row) for c in table.columns])
    # UTF-8 BOM so Excel opens Arabic/Unicode correctly on double-click.
    return buf.getvalue().encode("utf-8-sig")


def to_xlsx(table: ReportTable) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = (table.title or "Report")[:31]
    ws.sheet_view.rightToLeft = table.rtl
    bold = Font(bold=True)
    right = Alignment(horizontal="right")

    r = 1
    if table.title:
        ws.cell(row=r, column=1, value=table.title).font = Font(bold=True, size=14)
        r += 2
    for label, value in table.meta:
        ws.cell(row=r, column=1, value=label).font = bold
        ws.cell(row=r, column=2, value=value)
        r += 1
    if table.meta:
        r += 1

    header_row = r
    for ci, col in enumerate(table.columns, start=1):
        ws.cell(row=r, column=ci, value=col.label).font = bold
    r += 1

    def write_row(row: dict, emphasize: bool = False) -> None:
        nonlocal r
        for ci, col in enumerate(table.columns, start=1):
            value = row.get(col.key)
            if col.kind == "money" and value not in (None, ""):
                cell = ws.cell(row=r, column=ci, value=_major(value))
                cell.number_format = "#,##0.00"
            elif col.kind == "number" and value not in (None, ""):
                try:
                    cell = ws.cell(row=r, column=ci, value=float(value))
                except (TypeError, ValueError):
                    cell = ws.cell(row=r, column=ci, value=str(value))
            else:
                cell = ws.cell(row=r, column=ci, value="" if value in (None, "") else str(value))
            if col.align == "end":
                cell.alignment = right
            if emphasize:
                cell.font = bold
        r += 1

    for row in table.rows:
        write_row(row)
    for row in table.footer:
        write_row(row, emphasize=True)

    for ci, col in enumerate(table.columns, start=1):
        letter = ws.cell(row=header_row, column=ci).column_letter
        ws.column_dimensions[letter].width = max(len(col.label), 12) + 2

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def render_bytes(table: ReportTable, fmt: str) -> bytes:
    if fmt == "csv":
        return to_csv(table)
    if fmt == "xlsx":
        return to_xlsx(table)
    raise ValueError(f"unsupported export format {fmt!r}")


def export_response(table: ReportTable, fmt: str, filename_base: str) -> HttpResponse:
    payload = render_bytes(table, fmt)
    filename = f"{filename_base}-{dt.date.today().isoformat()}.{fmt}"
    resp = HttpResponse(payload, content_type=_CONTENT_TYPES[fmt])
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
