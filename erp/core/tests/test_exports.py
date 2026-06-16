"""Core tabular export renderers — CSV (UTF-8 BOM, minor→major money) and XLSX (RTL, numbers)."""
from __future__ import annotations

import io

from erp.core.exports import (
    Column,
    ReportTable,
    EXPORT_FORMATS,
    export_response,
    to_csv,
    to_xlsx,
)


def _table(rtl: bool = False) -> ReportTable:
    return ReportTable(
        title="Trial Balance",
        columns=[
            Column("code", "Code"),
            Column("name", "الحساب"),  # Arabic header must survive
            Column("debit", "Debit", kind="money", align="end"),
        ],
        rows=[
            {"code": "1000", "name": "Cash", "debit": 1140_00},
            {"code": "1100", "name": "AR", "debit": None},
        ],
        footer=[{"code": "", "name": "Total", "debit": 1140_00}],
        meta=[("Period", "2026-06")],
        rtl=rtl,
    )


def test_csv_has_bom_and_converts_minor_to_major():
    data = to_csv(_table())
    assert data.startswith(b"\xef\xbb\xbf")  # UTF-8 BOM so Excel detects encoding
    text = data.decode("utf-8-sig")
    assert "Trial Balance" in text
    assert "Period,2026-06" in text
    assert "1140.00" in text          # 1140_00 minor → 1140.00 major
    assert "الحساب" in text           # Arabic header preserved
    # A None money cell renders empty, not "0.00".
    assert ",AR," in text


def test_xlsx_roundtrips_numbers_and_sets_rtl():
    from openpyxl import load_workbook

    data = to_xlsx(_table(rtl=True))
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    assert ws.sheet_view.rightToLeft is True
    # Money cells are real numbers (major units), not strings — so Excel can sum them.
    values = {cell.value for row in ws.iter_rows() for cell in row}
    assert 1140.0 in values
    assert "الحساب" in values


def test_export_response_sets_download_headers():
    resp = export_response(_table(), "xlsx", "trial-balance")
    assert resp["Content-Type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment;" in resp["Content-Disposition"]
    assert resp["Content-Disposition"].endswith('.xlsx"')
    assert "csv" in EXPORT_FORMATS and "xlsx" in EXPORT_FORMATS
