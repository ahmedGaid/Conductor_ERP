"""Report endpoints serve CSV/XLSX when ``?export=`` is given, JSON otherwise (auth enforced)."""
from __future__ import annotations

import io

import pytest
from rest_framework.test import APIClient

from erp.identity.models import User

pytestmark = pytest.mark.django_db

XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _admin_client() -> APIClient:
    user = User.objects.create_user(username="exp_admin", password="Dev12345!")
    user.is_superuser = True
    user.save(update_fields=["is_superuser"])
    client = APIClient()
    client.force_authenticate(user=user)
    return client


def _books(client) -> None:
    for code, name, type_, is_cash in [
        ("1000", "Cash", "asset", True),
        ("3000", "Capital", "equity", False),
        ("4000", "Sales", "income", False),
        ("1100", "AR", "asset", False),
    ]:
        assert client.post("/api/accounting/accounts",
                           {"code": code, "name": name, "type": type_, "is_cash": is_cash},
                           format="json").status_code == 201
    assert client.post("/api/accounting/fiscal-years",
                       {"code": "2026", "start_date": "2026-01-01", "end_date": "2026-12-31"},
                       format="json").status_code == 201
    assert client.post("/api/accounting/periods",
                       {"fiscal_year_code": "2026", "code": "2026-06",
                        "start_date": "2026-06-01", "end_date": "2026-06-30"},
                       format="json").status_code == 201
    # One balanced entry so the reports have content.
    assert client.post("/api/accounting/journals",
                       {"date": "2026-06-15", "memo": "sale",
                        "lines": [{"account_code": "1100", "debit": 1140_00, "credit": 0},
                                  {"account_code": "4000", "debit": 0, "credit": 1140_00}]},
                       format="json").status_code == 201


def test_trial_balance_csv_export():
    client = _admin_client()
    _books(client)
    res = client.get("/api/accounting/reports/trial-balance?export=csv")
    assert res.status_code == 200
    assert res["Content-Type"].startswith("text/csv")
    assert "attachment;" in res["Content-Disposition"]
    body = res.content.decode("utf-8-sig")
    assert "Trial Balance" in body
    assert "1140.00" in body          # minor → major in the export


def test_trial_balance_xlsx_export_is_a_real_workbook():
    from openpyxl import load_workbook

    client = _admin_client()
    _books(client)
    res = client.get("/api/accounting/reports/trial-balance?export=xlsx&lang=ar")
    assert res.status_code == 200
    assert res["Content-Type"].startswith(XLSX_CT)
    wb = load_workbook(io.BytesIO(res.content))
    ws = wb.active
    assert ws.sheet_view.rightToLeft is True          # Arabic export → RTL sheet
    assert 1140.0 in {c.value for row in ws.iter_rows() for c in row}


def test_unknown_format_falls_back_to_json():
    client = _admin_client()
    _books(client)
    res = client.get("/api/accounting/reports/trial-balance?export=pdf")
    assert res.status_code == 200
    assert res["Content-Type"].startswith("application/json")
    assert res.data["data"]["is_balanced"] is True


def test_vat_return_export_smoke():
    client = _admin_client()
    _books(client)
    res = client.get("/api/accounting/reports/vat-return?from=2026-01-01&to=2026-12-31&export=csv")
    assert res.status_code == 200
    assert res["Content-Type"].startswith("text/csv")


def test_export_requires_auth():
    res = APIClient().get("/api/accounting/reports/trial-balance?export=csv")
    assert res.status_code in (401, 403)
